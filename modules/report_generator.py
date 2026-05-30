import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from collections import defaultdict


def _parse_num(val):
    try:
        return float(str(val).replace(',', '').strip())
    except:
        return 0


def _group_by_month(rows):
    """Group data rows by month. Expects row[0] to be a date string."""
    months = defaultdict(list)
    for row in rows:
        if not row:
            continue
        date_str = str(row[0]).strip()
        month_key = 'Unknown'
        # Handle YYYY-MM format from Dash IOT
        import re
        m = re.match(r'^(\d{4})-(\d{2})$', date_str)
        if m:
            from datetime import date
            try:
                dt = date(int(m.group(1)), int(m.group(2)), 1)
                month_key = dt.strftime('%B %Y')
            except:
                month_key = date_str
        else:
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                try:
                    dt = datetime.strptime(date_str, fmt)
                    month_key = dt.strftime('%B %Y')
                    break
                except:
                    pass
        months[month_key].append(row)
    return months


def _month_totals(rows):
    # Dash IOT columns: Date | Consumption | Production | Grid Import | Grid Export
    gen = exp = imp = cons = 0
    for row in rows:
        if len(row) > 1: cons += _parse_num(row[1])
        if len(row) > 2: gen  += _parse_num(row[2])
        if len(row) > 3: imp  += _parse_num(row[3])
        if len(row) > 4: exp  += _parse_num(row[4])
    return gen, exp, imp, cons


def build_excel_report(client, analysis_data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Solar Report'

    red_fill   = PatternFill('solid', fgColor='C00000')
    dark_fill  = PatternFill('solid', fgColor='1A1714')
    grey_fill  = PatternFill('solid', fgColor='2A2724')
    amber_fill = PatternFill('solid', fgColor='3A2800')
    white_font = Font(color='FFFFFF', bold=True, name='Calibri')
    normal_font = Font(color='FFFFFF', name='Calibri')

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Halfway Charge - Solar Report: {client['name']}"
    ws['A1'].font = Font(color='FFFFFF', bold=True, size=16, name='Calibri')
    ws['A1'].fill = dark_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    headers = ['Date', 'Consumption (kWh)', 'Production (kWh)', 'Grid Import (kWh)', 'Grid Export (kWh)', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = white_font
        cell.fill = red_fill
        cell.alignment = Alignment(horizontal='center')

    rows = analysis_data.get('rows', [])
    months = _group_by_month(rows)
    current_row = 3

    for month_name, month_rows in months.items():
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws.cell(row=current_row, column=1, value=month_name).font = Font(color='E8A020', bold=True, name='Calibri')
        ws.cell(row=current_row, column=1).fill = amber_fill
        current_row += 1

        for r_idx, row in enumerate(month_rows):
            fill = grey_fill if r_idx % 2 == 0 else dark_fill
            for c_idx, val in enumerate(row[:6], 1):
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                cell.font = normal_font
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

        gen, exp, imp, cons = _month_totals(month_rows)
        subtotal_data = [f'{month_name} Total', cons, gen, imp, exp, '']
        for c_idx, val in enumerate(subtotal_data, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=val)
            cell.font = white_font
            cell.fill = PatternFill('solid', fgColor='1E3A1E')
            cell.alignment = Alignment(horizontal='center')
        current_row += 2

    totals = analysis_data.get('totals', {})
    grand = ['GRAND TOTAL', totals.get('consumption',0), totals.get('generation',0),
             totals.get('import',0), totals.get('export',0), '']
    for c_idx, val in enumerate(grand, 1):
        cell = ws.cell(row=current_row, column=c_idx, value=val)
        cell.font = white_font
        cell.fill = red_fill
        cell.alignment = Alignment(horizontal='center')

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.sheet_view.showGridLines = False
    wb.save(output_path)
    return output_path


def build_html_report(client, analysis_data, bills=None):
    name    = client.get('name', 'Client')
    period  = analysis_data.get('period', '')
    rows    = analysis_data.get('rows', [])
    totals  = analysis_data.get('totals', {})
    savings = analysis_data.get('savings', {})

    total_gen  = totals.get('generation', 0)
    total_exp  = totals.get('export', 0)
    total_imp  = totals.get('import', 0)
    total_cons = totals.get('consumption', 0)
    self_consumed   = max(total_gen - total_exp, 0)
    self_sufficiency = round((self_consumed / total_cons) * 100) if total_cons else 0
    savings_amount   = savings.get('amount', 0)
    savings_pct      = savings.get('percent', 0)

    months = _group_by_month(rows)
    monthly_html = ''
    for month_name, month_rows in months.items():
        gen, exp, imp, cons = _month_totals(month_rows)
        self_c = max(gen - exp, 0)
        row_html = ''
        for i, row in enumerate(month_rows):
            cls = 'even' if i % 2 == 0 else 'odd'
            cells = ''.join(f'<td>{c}</td>' for c in row[:5])
            row_html += f'<tr class="{cls}">{cells}</tr>'

        monthly_html += f'''
        <div class="month-block">
          <div class="month-title">{month_name}</div>
          <div class="kpi-mini-grid">
            <div class="kpi-mini green"><span>{gen:,.0f}</span><small>Production kWh</small></div>
            <div class="kpi-mini amber"><span>{self_c:,.0f}</span><small>Self Consumed</small></div>
            <div class="kpi-mini red"><span>{imp:,.0f}</span><small>Grid Import</small></div>
            <div class="kpi-mini"><span>{cons:,.0f}</span><small>Consumption kWh</small></div>
          </div>
          <div class="table-wrap">
            <table>
              <thead><tr><th>Date</th><th>Consumption (kWh)</th><th>Production (kWh)</th><th>Grid Import (kWh)</th><th>Grid Export (kWh)</th></tr></thead>
              <tbody>{row_html}</tbody>
            </table>
          </div>
        </div>'''

    bills_html = ''
    if bills:
        bill_list = bills if isinstance(bills, list) else [bills]
        total_bill_kwh    = 0
        total_bill_amount = 0

        for i, bill in enumerate(bill_list, 1):
            fname = bill.get('filename', f'Meter {i}')
            bperiod = bill.get('period', '')
            bkwh    = bill.get('total_kwh') or 0
            bamount = bill.get('total_amount') or 0
            total_bill_kwh    += bkwh
            total_bill_amount += bamount

            rate_rows = ''
            for item in (bill.get('line_items') or []):
                if len(item) >= 2 and any(str(c).strip() for c in item):
                    is_total = any('total' in str(c).lower() for c in item)
                    style = ' class="rate-total"' if is_total else ''
                    cells = ''.join(f'<td>{c}</td>' for c in item)
                    rate_rows += f'<tr{style}>{cells}</tr>'

            bills_html += f'''
            <div class="meter-block">
              <div class="meter-title">Meter {i} - {fname}
                {'<span class="meter-pill">' + bperiod + '</span>' if bperiod else ''}
              </div>
              {'<div class="table-wrap"><table><tbody>' + rate_rows + '</tbody></table></div>' if rate_rows else '<p class="no-data">No line items extracted from this bill.</p>'}
            </div>'''

        if len(bill_list) > 1:
            bills_html += f'''
            <div class="combined-summary">
              <span>Combined Total</span>
              <span>{total_bill_kwh:,.0f} kWh &nbsp;|&nbsp; R {total_bill_amount:,.2f}</span>
            </div>'''

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Halfway Charge - {name} Solar Report</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Bricolage+Grotesque:opsz,wght@12..96,300;12..96,600;12..96,700&family=Hanken+Grotesk:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
  :root {{
    --bg: #0A0907; --surface: rgba(255,255,255,0.04); --line: rgba(255,255,255,0.08);
    --red: #C00000; --amber: #E8A020; --green: #22C55E;
    --font-d: 'Bricolage Grotesque', sans-serif; --font-b: 'Hanken Grotesk', sans-serif;
  }}
  body {{ background: var(--bg); color: #E8E4DF; font-family: var(--font-b); min-height: 100vh; position: relative; overflow-x: hidden; }}
  body::before {{ content: ''; position: fixed; inset: 0; background: radial-gradient(ellipse 80% 50% at 50% -10%, rgba(192,0,0,0.18) 0%, transparent 70%); pointer-events: none; z-index: 0; }}
  .wrap {{ position: relative; z-index: 1; max-width: 1100px; margin: 0 auto; padding: 48px 24px; }}
  .report-header {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 48px; }}
  .brand {{ display: flex; align-items: center; gap: 12px; }}
  .brand-name {{ font-family: var(--font-d); font-size: 1.2rem; font-weight: 600; color: #fff; }}
  .report-meta h1 {{ font-family: var(--font-d); font-size: 1.8rem; font-weight: 700; color: #fff; text-align: right; }}
  .report-meta p {{ color: rgba(232,228,223,0.45); font-size: 0.9rem; margin-top: 4px; text-align: right; }}
  .savings-hero {{ background: var(--surface); border: 1px solid var(--line); border-radius: 20px; padding: 48px; text-align: center; margin-bottom: 32px; backdrop-filter: blur(12px); }}
  .savings-hero .label {{ font-size: 0.8rem; letter-spacing: 0.12em; text-transform: uppercase; color: rgba(232,228,223,0.4); margin-bottom: 12px; }}
  .savings-hero .amount {{ font-family: var(--font-d); font-size: clamp(3rem, 8vw, 5.5rem); font-weight: 700; background: linear-gradient(135deg, #fff 0%, var(--red) 100%); -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text; line-height: 1; }}
  .savings-hero .sub {{ color: rgba(232,228,223,0.45); margin-top: 16px; }}
  .kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 48px; }}
  .kpi {{ background: var(--surface); border: 1px solid var(--line); border-radius: 16px; padding: 24px; backdrop-filter: blur(8px); }}
  .kpi .kpi-label {{ font-size: 0.72rem; letter-spacing: 0.1em; text-transform: uppercase; color: rgba(232,228,223,0.4); margin-bottom: 8px; }}
  .kpi .kpi-value {{ font-family: var(--font-d); font-size: 2rem; font-weight: 700; color: #fff; }}
  .kpi .kpi-unit {{ font-size: 0.8rem; color: rgba(232,228,223,0.4); margin-top: 4px; }}
  .kpi.red .kpi-value {{ color: var(--red); }} .kpi.green .kpi-value {{ color: var(--green); }} .kpi.amber .kpi-value {{ color: var(--amber); }}
  .section-title {{ font-family: var(--font-d); font-size: 1.3rem; font-weight: 700; color: #fff; margin: 48px 0 24px; padding-bottom: 12px; border-bottom: 1px solid var(--line); }}
  .month-block {{ margin-bottom: 40px; }}
  .month-title {{ font-family: var(--font-d); font-size: 1.05rem; font-weight: 600; color: var(--amber); margin-bottom: 16px; }}
  .kpi-mini-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-bottom: 16px; }}
  .kpi-mini {{ background: var(--surface); border: 1px solid var(--line); border-radius: 12px; padding: 16px; text-align: center; }}
  .kpi-mini span {{ display: block; font-family: var(--font-d); font-size: 1.35rem; font-weight: 700; color: #fff; }}
  .kpi-mini small {{ font-size: 0.72rem; color: rgba(232,228,223,0.4); letter-spacing: 0.06em; text-transform: uppercase; margin-top: 4px; display: block; }}
  .kpi-mini.green span {{ color: var(--green); }} .kpi-mini.amber span {{ color: var(--amber); }} .kpi-mini.red span {{ color: var(--red); }}
  .table-wrap {{ background: var(--surface); border: 1px solid var(--line); border-radius: 14px; overflow: hidden; margin-bottom: 16px; backdrop-filter: blur(8px); }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.875rem; }}
  thead th {{ background: rgba(192,0,0,0.18); color: rgba(232,228,223,0.6); font-weight: 500; letter-spacing: 0.06em; text-transform: uppercase; font-size: 0.72rem; padding: 12px 16px; text-align: left; border-bottom: 1px solid var(--line); }}
  tbody td {{ padding: 11px 16px; border-bottom: 1px solid var(--line); }}
  tbody tr:last-child td {{ border-bottom: none; }}
  tbody tr.even td {{ background: rgba(255,255,255,0.02); }}
  tbody tr.rate-total td {{ background: #C00000 !important; color: #fff !important; font-weight: 600 !important; }}
  .meter-block {{ margin-bottom: 32px; }}
  .meter-title {{ font-family: var(--font-d); font-size: 1rem; font-weight: 600; color: #fff; margin-bottom: 14px; display: flex; align-items: center; gap: 10px; }}
  .meter-pill {{ background: rgba(192,0,0,0.2); color: var(--red); border-radius: 20px; padding: 3px 10px; font-size: 0.75rem; font-weight: 500; }}
  .combined-summary {{ display: flex; justify-content: space-between; align-items: center; background: rgba(192,0,0,0.12); border: 1px solid rgba(192,0,0,0.3); border-radius: 12px; padding: 16px 20px; font-weight: 600; color: #fff; margin-top: 8px; }}
  .no-data {{ color: rgba(232,228,223,0.3); font-size: 0.85rem; padding: 12px 0; }}
  .report-footer {{ text-align: center; margin-top: 64px; color: rgba(232,228,223,0.2); font-size: 0.8rem; }}
</style>
</head>
<body>
<div class="wrap">
  <div class="report-header">
    <div class="brand">
      <svg width="32" height="32" viewBox="0 0 32 32" fill="none"><polygon points="18,2 7,18 16,18 14,30 25,14 16,14" fill="#C00000"/></svg>
      <span class="brand-name">Halfway Charge</span>
    </div>
    <div class="report-meta">
      <h1>{name}</h1>
      <p>Solar Performance Report - {period}</p>
      <p>Generated {datetime.now().strftime('%d %B %Y')}</p>
    </div>
  </div>
  <div class="savings-hero">
    <div class="label">Estimated Savings This Period</div>
    <div class="amount">R {savings_amount:,.0f}</div>
    <div class="sub">{savings_pct:.1f}% reduction vs grid-only baseline</div>
  </div>
  <div class="kpi-grid">
    <div class="kpi green"><div class="kpi-label">Total Production</div><div class="kpi-value">{total_gen:,.0f}</div><div class="kpi-unit">kWh</div></div>
    <div class="kpi amber"><div class="kpi-label">Self Consumed</div><div class="kpi-value">{self_consumed:,.0f}</div><div class="kpi-unit">kWh</div></div>
    <div class="kpi red"><div class="kpi-label">Grid Import</div><div class="kpi-value">{total_imp:,.0f}</div><div class="kpi-unit">kWh</div></div>
    <div class="kpi"><div class="kpi-label">Self-Sufficiency</div><div class="kpi-value">{self_sufficiency}</div><div class="kpi-unit">%</div></div>
  </div>
  {('<div class="section-title">Solar Data by Month</div>' + monthly_html) if monthly_html else ''}
  {('<div class="section-title">Eskom Bill Breakdown</div>' + bills_html) if bills_html else ''}
  <div class="report-footer">
    <p>Halfway Charge Analytics - Confidential</p>
    <p>Generated automatically from Dash IOT monitoring data.</p>
  </div>
</div>
</body>
</html>"""
    return html
