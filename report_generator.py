"""
Generates Excel report and shareable HTML report for a solar analysis.
"""
import os
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RED = "C00000"
LIGHT_RED = "FFF2F2"
WHITE = "FFFFFF"
GREEN = "1A7A4A"
LIGHT_GREEN = "E8F5EE"
GREY = "F2F2F2"

thin = Side(style="thin", color="CCCCCC")
brd = Border(top=thin, bottom=thin, left=thin, right=thin)
hdr_fill = PatternFill("solid", fgColor=RED)
hdr_font = Font(bold=True, color=WHITE, name="Arial", size=11)
alt_fill = PatternFill("solid", fgColor=LIGHT_RED)


def build_excel(analysis, client, rate_components, daily_rows, output_dir):
    wb = Workbook()

    # ── Sheet 1: Raw Daily Data ───────────────────────────────
    ws1 = wb.active
    ws1.title = "Raw Data"
    headers = ["Date", "Consumption [kWh]", "Production [kWh]", "Grid Import [kWh]", "Grid Export [kWh]"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = brd

    for r_i, row in enumerate(daily_rows, 2):
        vals = [row["date"], row["consumption"], row["production"], row["grid_import"], row["grid_export"]]
        for c_i, v in enumerate(vals, 1):
            cell = ws1.cell(row=r_i, column=c_i, value=v)
            cell.font = Font(name="Arial", size=10); cell.border = brd
            cell.alignment = Alignment(horizontal="center")
            if r_i % 2 == 0: cell.fill = alt_fill

    n = len(daily_rows) + 2
    ws1.cell(row=n, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=10)
    ws1.cell(row=n, column=1).fill = PatternFill("solid", fgColor="FFE0E0")
    for c in range(2, 6):
        cell = ws1.cell(row=n, column=c, value=f"=SUM({get_column_letter(c)}2:{get_column_letter(c)}{n-1})")
        cell.font = Font(bold=True, name="Arial", size=10, color="0000FF")
        cell.fill = PatternFill("solid", fgColor="FFE0E0"); cell.border = brd
        cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal="center")

    ws1.column_dimensions["A"].width = 14
    for col in ["B","C","D","E"]: ws1.column_dimensions[col].width = 22

    # ── Sheet 2: Analysis ─────────────────────────────────────
    ws2 = wb.create_sheet("Analysis")

    def sec(row, text):
        for c in range(1, 4):
            ws2.cell(row=row, column=c).fill = hdr_fill
            ws2.cell(row=row, column=c).border = brd
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws2.cell(row=row, column=1, value=text)
        cell.font = hdr_font; cell.alignment = Alignment(horizontal="left", indent=1)

    def row_(r, label, value, fmt=None, note=None, bold=False, color="000000"):
        lc = ws2.cell(row=r, column=1, value=label)
        lc.font = Font(name="Arial", size=10); lc.border = brd
        lc.alignment = Alignment(horizontal="left", indent=1)
        vc = ws2.cell(row=r, column=2, value=value)
        vc.font = Font(name="Arial", size=10, bold=bold, color=color)
        vc.border = brd; vc.alignment = Alignment(horizontal="right")
        if fmt: vc.number_format = fmt
        nc = ws2.cell(row=r, column=3); nc.border = brd
        if note: nc.value = note; nc.font = Font(name="Arial", size=9, italic=True, color="666666")

    r = 1
    ws2.cell(row=r, column=1, value=f"SOLAR ANALYSIS — {client['name'].upper()}").font = Font(bold=True, name="Arial", size=14, color=RED)
    ws2.merge_cells(f"A{r}:C{r}"); r += 1
    ws2.cell(row=r, column=1, value=f"Period: {analysis['bill_period_from']} to {analysis['bill_period_to']}").font = Font(name="Arial", size=10, italic=True)
    ws2.merge_cells(f"A{r}:C{r}"); r += 2

    sec(r, "ELECTRICITY BILL"); r += 1
    row_(r, "Total Grid kWh Billed", analysis["bill_kwh"], '#,##0.00'); r += 1
    row_(r, "Total Charges (excl VAT)", analysis["bill_excl_vat"], '"R" #,##0.00'); r += 1
    row_(r, "Total Amount Due (incl VAT)", analysis["bill_incl_vat"], '"R" #,##0.00', bold=True); r += 2

    sec(r, "DASHBOARD DATA"); r += 1
    row_(r, "Total Consumption", analysis["dash_consumption"], '#,##0.00 "kWh"'); r += 1
    row_(r, "Solar Production", analysis["dash_production"], '#,##0.00 "kWh"'); r += 1
    row_(r, "Grid Import (Dashboard)", analysis["dash_grid_import"], '#,##0.00 "kWh"'); r += 1
    row_(r, "Self-Consumed Solar", analysis["dash_production"], '#,##0.00 "kWh"', "Production – Export"); r += 2

    sec(r, "VALIDATION"); r += 1
    row_(r, "Dashboard Grid Import", analysis["dash_grid_import"], '#,##0.00 "kWh"'); r += 1
    row_(r, "Eskom Bill kWh", analysis["bill_kwh"], '#,##0.00 "kWh"'); r += 1
    variance_kwh = round(analysis["dash_grid_import"] - analysis["bill_kwh"], 2)
    row_(r, "Variance (kWh)", variance_kwh, '#,##0.00 "kWh"'); r += 1
    row_(r, "Variance (%)", round(analysis["variance_pct"], 2)/100, '0.00%', "✓ Validated" if analysis["variance_pct"] < 3 else "⚠ Review", color="008000"); r += 2

    sec(r, "SAVINGS"); r += 1
    row_(r, "Combined kWh Savings Rate", analysis["combined_rate"], '"R" #,##0.0000 "/kWh"'); r += 1
    row_(r, "Estimated Savings (excl VAT)", analysis["savings_excl_vat"], '"R" #,##0.00', bold=True, color="008000"); r += 1
    row_(r, "Estimated Savings (incl VAT)", analysis["savings_incl_vat"], '"R" #,##0.00', bold=True, color="008000"); r += 2

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 30

    # ── Sheet 3: Rate Breakdown ───────────────────────────────
    ws3 = wb.create_sheet("Rate Breakdown")
    ws3.cell(row=1, column=1, value="PER-kWh SAVINGS RATE BREAKDOWN").font = Font(bold=True, name="Arial", size=13, color=RED)
    ws3.merge_cells("A1:D1")
    for c, h in enumerate(["Charge Type", "Applies To", "Rate (R/kWh)", "Notes"], 1):
        cell = ws3.cell(row=2, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = brd
    for r_i, comp in enumerate(rate_components, 3):
        vals = [comp["name"], comp.get("applies_to",""), round(comp["rate"],4), ""]
        for c_i, v in enumerate(vals, 1):
            cell = ws3.cell(row=r_i, column=c_i, value=v)
            cell.font = Font(name="Arial", size=10); cell.border = brd
            if r_i % 2 == 0: cell.fill = alt_fill
    total_row = len(rate_components) + 3
    total_rate = analysis["combined_rate"]
    for c in range(1, 5):
        cell = ws3.cell(row=total_row, column=c)
        cell.fill = hdr_fill; cell.border = brd
    ws3.cell(row=total_row, column=1, value="COMBINED SAVINGS RATE").font = Font(bold=True, color=WHITE, name="Arial")
    ws3.cell(row=total_row, column=3, value=total_rate).font = Font(bold=True, color=WHITE, name="Arial")
    ws3.column_dimensions["A"].width = 30; ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 16; ws3.column_dimensions["D"].width = 36

    safe_name = client['name'].replace(' ','_').replace('/','').replace('\\','').strip('_')
    fname = f"{safe_name}_{analysis['bill_period_from'][:7]}_Analysis.xlsx"
    os.makedirs(output_dir, exist_ok=True)
    fpath = os.path.join(output_dir, fname)
    wb.save(fpath)
    return fpath


def build_html_report(analysis, client, rate_components, token):
    """Returns HTML string for the shareable report page."""
    variance_ok = analysis["variance_pct"] < 3
    solar_offset = round(analysis["dash_production"] / analysis["dash_consumption"] * 100, 1) if analysis["dash_consumption"] else 0
    grid_dep = round(100 - solar_offset, 1)
    variance_color = "#2ECC71" if variance_ok else "#FF6B6B"
    variance_icon = "✓" if variance_ok else "⚠"

    rate_rows = ""
    for comp in rate_components:
        rate_rows += f"""
        <tr>
          <td>{comp['name']}</td>
          <td style="color:#9B958A;">{comp.get('applies_to','')}</td>
          <td style="text-align:right;font-family:'Bricolage Grotesque',sans-serif;font-weight:700;color:#F4A300;">R {comp['rate']:.4f}</td>
        </tr>"""

    # Bar chart widths for savings visualization
    max_val = max(analysis['savings_incl_vat'], analysis['dash_production'] * analysis['combined_rate'] * 1.15, 1)
    bar_w = min(100, round(analysis['savings_incl_vat'] / max_val * 100))

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Solar Report — {client['name']}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Bricolage+Grotesque:opsz,wght@12..96,500;12..96,700;12..96,800&family=Hanken+Grotesk:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
  :root {{
    --bg:#0A0907; --bg2:#13110C; --surface:rgba(255,255,255,.045); --surface2:rgba(255,255,255,.08);
    --line:rgba(255,255,255,.10); --line2:rgba(255,255,255,.18);
    --text:#F7F4EC; --muted:#9B958A;
    --red:#C00000; --red-glow:rgba(192,0,0,.35);
    --amber:#F4A300; --green:#2ECC71;
    --font-d:'Bricolage Grotesque',sans-serif; --font-b:'Hanken Grotesk',sans-serif;
  }}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:var(--font-b);background:var(--bg);color:var(--text);min-height:100vh;}}
  body::before{{content:"";position:fixed;inset:0;z-index:0;pointer-events:none;
    background:radial-gradient(900px 600px at 85% -5%,rgba(192,0,0,.12),transparent 60%),
               radial-gradient(700px 700px at -10% 110%,rgba(244,163,0,.08),transparent 60%);}}
  .grain{{position:fixed;inset:0;z-index:0;pointer-events:none;opacity:.04;mix-blend-mode:overlay;
    background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='120' height='120'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='.85' numOctaves='3'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)'/%3E%3C/svg%3E");}}
  .page{{position:relative;z-index:1;max-width:960px;margin:0 auto;padding:clamp(20px,3vw,40px);}}

  /* HEADER */
  .header{{display:flex;align-items:center;gap:20px;padding-bottom:28px;border-bottom:1px solid var(--line);margin-bottom:28px;flex-wrap:wrap;}}
  .logo-wrap{{background:white;border-radius:12px;padding:8px 12px;flex-shrink:0;}}
  .logo-wrap img{{height:52px;display:block;}}
  .header-text h1{{font-family:var(--font-d);font-weight:800;font-size:clamp(18px,3vw,26px);letter-spacing:-.02em;}}
  .header-text p{{color:var(--muted);font-size:13px;margin-top:5px;line-height:1.5;}}
  .header-badge{{margin-left:auto;background:var(--surface);border:1px solid var(--line2);border-radius:12px;padding:10px 16px;text-align:right;}}
  .header-badge .b-label{{font-size:10px;letter-spacing:.2em;text-transform:uppercase;color:var(--muted);}}
  .header-badge .b-val{{font-family:var(--font-d);font-weight:800;font-size:18px;color:var(--text);margin-top:2px;}}

  /* KPI BAR */
  .kpi-grid{{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:24px;}}
  .kpi{{background:var(--surface);border:1px solid var(--line);border-radius:14px;padding:16px 12px;text-align:center;}}
  .kpi .kval{{font-family:var(--font-d);font-weight:800;font-size:clamp(18px,2.5vw,26px);letter-spacing:-.02em;display:block;}}
  .kpi .kunit{{font-size:11px;color:var(--muted);display:block;margin-top:1px;}}
  .kpi .klbl{{font-size:10px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;color:var(--muted);display:block;margin-top:5px;}}
  .kpi.k-red{{border-color:rgba(192,0,0,.4);background:rgba(192,0,0,.12);}} .kpi.k-red .kval{{color:var(--red);}}
  .kpi.k-amber{{border-color:rgba(244,163,0,.4);background:rgba(244,163,0,.10);}} .kpi.k-amber .kval{{color:var(--amber);}}
  .kpi.k-blue{{border-color:rgba(31,78,121,.5);background:rgba(31,78,121,.18);}} .kpi.k-blue .kval{{color:#5B9BD5;}}
  .kpi.k-green{{border-color:rgba(46,204,113,.3);background:rgba(46,204,113,.10);}} .kpi.k-green .kval{{color:var(--green);}}
  .kpi.k-savings{{border-color:rgba(192,0,0,.5);background:linear-gradient(135deg,rgba(192,0,0,.18),rgba(244,163,0,.08));}}
  .kpi.k-savings .kval{{font-size:clamp(16px,2.2vw,22px);color:var(--text);}}

  /* SECTION */
  .section{{margin-bottom:20px;}}
  .section-head{{font-size:11px;font-weight:700;letter-spacing:.2em;text-transform:uppercase;color:var(--red);margin-bottom:12px;display:flex;align-items:center;gap:10px;}}
  .section-head::after{{content:"";flex:1;height:1px;background:var(--line);}}

  /* TABLE */
  .card-table{{background:var(--surface);border:1px solid var(--line);border-radius:16px;overflow:hidden;}}
  table{{width:100%;border-collapse:collapse;font-size:13px;}}
  thead th{{padding:10px 14px;font-size:10px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;color:var(--muted);border-bottom:1px solid var(--line2);text-align:left;}}
  tbody tr{{border-bottom:1px solid var(--line);}}
  tbody tr:last-child{{border-bottom:none;}}
  tbody td{{padding:11px 14px;vertical-align:middle;}}
  .row-highlight td{{background:rgba(46,204,113,.07)!important;}}
  .row-total td{{background:rgba(192,0,0,.12)!important;border-top:1px solid rgba(192,0,0,.3);}}
  .row-total td:first-child{{font-weight:700;color:var(--text);}}
  .row-total td:last-child{{font-family:var(--font-d);font-weight:800;font-size:16px;color:var(--red);}}

  /* VARIANCE BOX */
  .variance-box{{border-radius:12px;padding:12px 16px;font-size:13px;font-weight:600;margin-top:12px;display:flex;align-items:center;gap:10px;}}
  .variance-ok{{background:rgba(46,204,113,.1);border:1px solid rgba(46,204,113,.3);color:var(--green);}}
  .variance-warn{{background:rgba(192,0,0,.1);border:1px solid rgba(192,0,0,.3);color:#FF6B6B;}}

  /* SAVINGS HERO */
  .savings-hero{{background:linear-gradient(135deg,rgba(192,0,0,.15),rgba(244,163,0,.08));border:1px solid rgba(192,0,0,.3);border-radius:18px;padding:28px;display:flex;align-items:center;gap:28px;flex-wrap:wrap;margin-bottom:20px;}}
  .savings-num{{font-family:var(--font-d);font-weight:800;letter-spacing:-.04em;font-size:clamp(48px,8vw,80px);line-height:.85;background:linear-gradient(120deg,var(--red),var(--amber));-webkit-background-clip:text;background-clip:text;color:transparent;}}
  .savings-sub{{color:var(--muted);font-size:14px;margin-top:10px;line-height:1.5;}}
  .savings-sub b{{color:var(--text);}}

  /* NOTE */
  .note{{color:var(--muted);font-size:11.5px;margin-top:10px;line-height:1.6;padding:10px 14px;background:var(--surface);border-radius:10px;border:1px solid var(--line);}}

  /* FOOTER */
  .footer{{margin-top:32px;padding-top:20px;border-top:1px solid var(--line);text-align:center;color:var(--muted);font-size:11px;line-height:1.6;}}

  /* PRINT */
  @media print{{
    body::before,.grain{{display:none;}}
    body{{background:white;color:#111;}}
    .page{{padding:20px;}}
    .savings-num{{color:var(--red)!important;-webkit-text-fill-color:var(--red)!important;}}
  }}
  @media(max-width:640px){{
    .kpi-grid{{grid-template-columns:repeat(2,1fr);}}
    .kpi.k-savings{{grid-column:span 2;}}
  }}
</style>
</head>
<body>
<div class="grain"></div>
<div class="page">

  <!-- HEADER -->
  <div class="header">
    <div class="logo-wrap"><img src="/static/logo.jpg" alt="Halfway"></div>
    <div class="header-text">
      <h1>Solar Performance Report</h1>
      <p>{client['name']} &nbsp;·&nbsp; {analysis['bill_period_from']} to {analysis['bill_period_to']}<br>Prepared by Halfway Charge Analytics</p>
    </div>
    <div class="header-badge">
      <div class="b-label">Period</div>
      <div class="b-val">{analysis['bill_period_from'][:7]}</div>
    </div>
  </div>

  <!-- KPI BAR -->
  <div class="kpi-grid">
    <div class="kpi k-red"><span class="kval">{analysis['dash_consumption']:,.0f}</span><span class="kunit">kWh</span><span class="klbl">Consumption</span></div>
    <div class="kpi k-amber"><span class="kval">{analysis['dash_production']:,.0f}</span><span class="kunit">kWh</span><span class="klbl">Solar Production</span></div>
    <div class="kpi k-blue"><span class="kval">{analysis['dash_grid_import']:,.0f}</span><span class="kunit">kWh</span><span class="klbl">Grid Import</span></div>
    <div class="kpi k-green"><span class="kval">{solar_offset:.1f}%</span><span class="kunit">of consumption</span><span class="klbl">Solar Offset</span></div>
    <div class="kpi k-savings"><span class="kval">R {analysis['savings_incl_vat']:,.0f}</span><span class="kunit">incl. VAT</span><span class="klbl">Est. Savings</span></div>
  </div>

  <!-- SAVINGS HERO -->
  <div class="savings-hero">
    <div>
      <div style="font-size:11px;letter-spacing:.2em;text-transform:uppercase;color:var(--red);font-weight:700;margin-bottom:8px;">Estimated Monthly Savings</div>
      <div class="savings-num">R {analysis['savings_incl_vat']:,.0f}</div>
      <div class="savings-sub">Solar offset of <b>{solar_offset:.1f}%</b> saved <b>R {analysis['savings_excl_vat']:,.2f} excl VAT</b> at R {analysis['combined_rate']:.4f}/kWh combined rate.</div>
    </div>
    <div style="flex:1;min-width:200px;">
      <div style="font-size:11px;color:var(--muted);margin-bottom:10px;font-weight:600;">Solar vs Grid share</div>
      <div style="height:10px;background:var(--line2);border-radius:99px;overflow:hidden;margin-bottom:6px;">
        <div style="height:100%;width:{solar_offset:.1f}%;background:linear-gradient(90deg,var(--amber),var(--red));border-radius:99px;transition:width 1s;"></div>
      </div>
      <div style="display:flex;justify-content:space-between;font-size:11px;color:var(--muted);">
        <span style="color:var(--amber);font-weight:700;">☀ {solar_offset:.1f}% solar</span>
        <span>{grid_dep:.1f}% grid</span>
      </div>
    </div>
  </div>

  <!-- GRID IMPORT VALIDATION -->
  <div class="section">
    <div class="section-head">Grid Import Validation</div>
    <div class="card-table">
      <table>
        <thead><tr><th></th><th>Dashboard (Dash IOT)</th><th>Eskom Bill</th><th>Notes</th></tr></thead>
        <tbody>
          <tr><td>Grid Import (kWh)</td><td>{analysis['dash_grid_import']:,.1f} kWh</td><td>{analysis['bill_kwh']:,.2f} kWh</td><td style="color:var(--muted);">Bill period match</td></tr>
          <tr class="row-highlight">
            <td style="font-weight:700;">Variance</td>
            <td style="color:{variance_color};font-weight:700;">{analysis['dash_grid_import'] - analysis['bill_kwh']:,.1f} kWh</td>
            <td style="color:{variance_color};font-weight:700;">{analysis['variance_pct']:.2f}%</td>
            <td style="color:{variance_color};font-weight:700;">{variance_icon} {"Validated" if variance_ok else "Review required"}</td>
          </tr>
        </tbody>
      </table>
    </div>
    <div class="variance-box {'variance-ok' if variance_ok else 'variance-warn'}">
      <span style="font-size:18px;">{"✓" if variance_ok else "⚠"}</span>
      <span>{"Dashboard and Eskom meter readings are aligned. Variance of " + f"{analysis['variance_pct']:.2f}% is within the 3% threshold." if variance_ok else f"Variance of {analysis['variance_pct']:.2f}% exceeds 3% — please review meter readings before issuing report."}</span>
    </div>
  </div>

  <!-- SOLAR PERFORMANCE -->
  <div class="section">
    <div class="section-head">Solar Performance Detail</div>
    <div class="card-table">
      <table>
        <thead><tr><th></th><th>Dashboard (Dash IOT)</th><th>Eskom Bill</th><th>Notes</th></tr></thead>
        <tbody>
          <tr><td>Total Site Consumption</td><td>{analysis['dash_consumption']:,.1f} kWh</td><td>—</td><td></td></tr>
          <tr><td>Solar Production</td><td style="color:var(--amber);font-weight:600;">{analysis['dash_production']:,.1f} kWh</td><td>—</td><td></td></tr>
          <tr><td>Self-Consumed Solar</td><td>{analysis['dash_production']:,.1f} kWh</td><td>—</td><td style="color:var(--muted);">No grid export</td></tr>
          <tr><td>Solar Offset</td><td style="color:var(--green);font-weight:700;">{solar_offset:.1f}%</td><td>—</td><td style="color:var(--muted);">% of load from solar</td></tr>
          <tr><td>Grid Dependency</td><td>{grid_dep:.1f}%</td><td>—</td><td style="color:var(--muted);">% of load from grid</td></tr>
        </tbody>
      </table>
    </div>
  </div>

  <!-- SAVINGS RATE -->
  <div class="section">
    <div class="section-head">Savings Rate Breakdown</div>
    <div class="card-table">
      <table>
        <thead><tr><th>Per-kWh Charge</th><th>Applies To</th><th style="text-align:right;">Rate (R/kWh)</th></tr></thead>
        <tbody>
          {rate_rows}
          <tr class="row-total">
            <td>Combined Savings Rate (excl VAT)</td><td></td>
            <td style="text-align:right;">R {analysis['combined_rate']:.4f}</td>
          </tr>
        </tbody>
      </table>
    </div>
    <div class="note">Fixed charges (administration, network capacity, generator capacity, service charge) are excluded — these are billed per day or per kVA and are not reduced by solar generation.</div>
  </div>

  <!-- FINANCIAL SUMMARY -->
  <div class="section">
    <div class="section-head">Financial Summary</div>
    <div class="card-table">
      <table>
        <thead><tr><th></th><th>Value</th><th>Notes</th></tr></thead>
        <tbody>
          <tr><td>Total Bill (excl VAT)</td><td>R {analysis['bill_excl_vat']:,.2f}</td><td style="color:var(--muted);">From Eskom bill</td></tr>
          <tr><td>Total Bill (incl VAT)</td><td>R {analysis['bill_incl_vat']:,.2f}</td><td style="color:var(--muted);">From Eskom bill</td></tr>
          <tr><td>Combined kWh Savings Rate</td><td style="color:var(--amber);font-weight:700;">R {analysis['combined_rate']:.4f} / kWh</td><td style="color:var(--muted);">excl VAT</td></tr>
          <tr><td>Est. Solar Savings (excl VAT)</td><td style="color:var(--green);font-weight:700;">R {analysis['savings_excl_vat']:,.2f}</td><td></td></tr>
          <tr class="row-highlight"><td style="font-weight:700;">Est. Solar Savings (incl VAT)</td><td style="font-family:'Bricolage Grotesque',sans-serif;font-weight:800;font-size:18px;color:var(--green);">R {analysis['savings_incl_vat']:,.2f}</td><td style="color:var(--green);font-weight:700;">★ Benefit to client</td></tr>
        </tbody>
      </table>
    </div>
  </div>

  <div class="footer">
    Report generated by Halfway Charge Analytics &nbsp;·&nbsp; {analysis['bill_period_from']} – {analysis['bill_period_to']} &nbsp;·&nbsp; Confidential
  </div>

</div>
</body>
</html>"""
    return html
